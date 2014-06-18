import csv
from openpyxl import Workbook
from openpyxl.cell import get_column_letter
import glob
import os

wb = Workbook()

i=0
for file in glob.glob("*.csv"):
    f = open(file)
    print(file)
    csv.register_dialect('colons', delimiter=',')
    reader = csv.reader(f, dialect='colons')
    wb.create_sheet()
    ws = wb.worksheets[i]
    ws.title = file

    for row_index, row in enumerate(reader):
        for column_index, cell in enumerate(row):
            column_letter = get_column_letter((column_index + 1))
            ws.cell('%s%s'%(column_letter, (row_index + 1))).value = cell

    i+=1

dest_filename = "Copa2014_Estatisticas.xls"
wb.save(filename = dest_filename)


